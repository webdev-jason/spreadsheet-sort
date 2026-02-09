import openpyxl
import sys
import os
import shutil
import warnings

# We still suppress warnings to keep the UI clean
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def process_spreadsheet_with_styles(file_path, target_lot):
    try:
        print(f"Opening master file: {os.path.basename(file_path)}...")
        
        # 1. Create the output path
        output_filename = f"Lot_{target_lot}_Formatted.xlsx"
        output_path = os.path.join(os.path.dirname(file_path), output_filename)
        
        # 2. Copy the entire file first to preserve ALL formatting and styles
        # This is the "Template" approach
        shutil.copy2(file_path, output_path)
        
        # 3. Load the COPIED workbook so we can edit it
        # data_only=False ensures we keep formulas; keep_vba=True keeps macros
        wb = openpyxl.load_workbook(output_path, data_only=False)
        
        sheets_to_delete = []
        match_count = 0
        total_sheets = len(wb.sheetnames)

        # 4. Iterate and find which sheets to keep
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found_in_sheet = False
            
            # Check Column C (Column 3) for the lot number
            # We iterate through the rows in Column C
            for cell in ws['C']:
                if cell.value and str(target_lot).lower() in str(cell.value).lower():
                    found_in_sheet = True
                    break
            
            if found_in_sheet:
                match_count += 1
            else:
                sheets_to_delete.append(sheet_name)
        
        # 5. Remove the non-matching sheets
        if match_count > 0:
            for name in sheets_to_delete:
                del wb[name]
            
            wb.save(output_path)
            print(f"SUCCESS: Extracted {match_count} sheets with full formatting to:\n{output_filename}")
        else:
            # Clean up the copy if no matches were found
            wb.close()
            if os.path.exists(output_path):
                os.remove(output_path)
            print(f"NOT FOUND: Lot {target_lot} not found in any of the {total_sheets} sheets.")

    except Exception as e:
        print(f"CRITICAL ERROR: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 2:
        file_path = sys.argv[1]
        target_lot = sys.argv[2]
        process_spreadsheet_with_styles(file_path, target_lot)
    else:
        print("ERROR: Missing arguments.")